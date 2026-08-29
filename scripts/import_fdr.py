#!/usr/bin/env python3
"""
Import the hand-built FDR workbook (data/ucl_fdr_source.xlsx) into
data/ucl_fdr.json, and fold it into the per-matchday projections.

    python3 scripts/import_fdr.py
    python3 scripts/import_fdr.py --check     # report only, write nothing

The workbook encodes difficulty as CELL FILL COLOUR, not values:

    6AA84F dark green  1  Very easy        (UCL debutants)
    93C47D light green 2  Easy             (most Pot 3-4 teams)
    CCCCCC grey        3  Anyone's game    (50/50)
    EA9999 light red   4  Hard             (Pot 1-2 at home)
    E06666 dark red    5  Very hard        (Pot 1-2 away)

Rows are identified by their EIGHT FIXTURES matched against the validated
calendar, not by the row label. The source labels two rows wrongly — the row
marked SLB holds Slavia Praha's fixtures and the one marked SLP holds Slovan
Bratislava's, the opposite of the workbook's own legend — so trusting labels
would silently swap two clubs' difficulty. The fixtures are unambiguous.
"""
import argparse
import json
import os
import re
import sys

import openpyxl

HERE = os.path.dirname(__file__)
sys.path.insert(0, os.path.normpath(os.path.join(HERE, "..")))
DATA = os.path.normpath(os.path.join(HERE, "..", "data"))

import data.team_stats as ts                     # noqa: E402

SRC = os.path.join(DATA, "ucl_fdr_source.xlsx")
OUT = os.path.join(DATA, "ucl_fdr.json")
PROJ = os.path.join(DATA, "ucl_md_projections.json")

COLOUR_TO_FDR = {
    "6AA84F": 1,   # very easy
    "93C47D": 2,   # easy
    "CCCCCC": 3,   # anyone's game
    "EA9999": 4,   # hard
    "E06666": 5,   # very hard
}

# The workbook's in-cell abbreviations, per its own legend, mapped to our codes.
CELL_ABBR = {
    "BAR": "BAR", "BET": "BET", "BOD": "BOD", "BRU": "CLB", "FCB": "BAY",
    "BAY": "BAY", "FCP": "POR", "FEN": "FEN", "FEY": "FEY", "MCI": "MCI",
    "MUN": "MUN", "RCL": "LEN", "ROM": "ROM", "SLB": "SLB", "SLP": "SLA",
    "STU": "STU", "VIK": "VIK", "LIV": "LIV", "RMA": "RMA", "INT": "INT",
    "BVB": "DOR", "ARS": "ARS", "PSG": "PSG", "NAP": "NAP", "SPO": "SPO",
    "AVL": "AVL", "ATM": "ATM", "GAL": "GAL", "PSV": "PSV", "RBL": "RBL",
    "SHK": "SHK", "VIL": "VIL", "LIL": "LIL", "COM": "COM", "AEK": "AEK",
    "LAS": "LSK", "SAB": "SAB",
}


def cell_fill(cell) -> str:
    try:
        return (cell.fill.fgColor.rgb or "")[-6:].upper()
    except Exception:
        return ""


def read_rows(ws):
    """Yield (row_label, [(opp_code, is_home, fdr) x8]) for each team row."""
    for r in range(1, ws.max_row + 1):
        label = ws.cell(r, 2).value
        if not label or str(label).strip() in ("Team", "Ranking"):
            continue
        label = str(label).strip()
        if label not in CELL_ABBR:
            continue
        games, ok = [], True
        for md in range(1, 9):
            cell = ws.cell(r, 2 + md)
            m = re.match(r"([A-Z]{3})\s*\((H|A)\)", str(cell.value or "").strip())
            if not m:
                ok = False
                break
            opp = CELL_ABBR.get(m.group(1))
            fdr = COLOUR_TO_FDR.get(cell_fill(cell))
            if not opp or not fdr:
                ok = False
                break
            games.append((opp, m.group(2) == "H", fdr))
        if ok and len(games) == 8:
            yield label, games


def identify(games, allow_wrong: int = 1) -> tuple:
    """Which club's calendar matches these eight (opponent, venue) pairs?

    Tolerates up to `allow_wrong` mismatched cells so a single typo in the
    source doesn't discard a whole club, and reports which cells disagreed —
    the colour (the actual difficulty rating) is still usable.
    Returns (matching_codes, mismatches_for_the_single_best_match).
    """
    scored = []
    for code in ts.TEAM_NAMES:
        wrong = [
            (md, opp, ts.get_md_fixture(md, code))
            for md, (opp, home, _) in enumerate(games, start=1)
            if ts.get_md_fixture(md, code) != opp or ts.is_home(md, code) != home
        ]
        if len(wrong) <= allow_wrong:
            scored.append((len(wrong), code, wrong))
    if not scored:
        return [], []
    best = min(s[0] for s in scored)
    hits = [(c, w) for n, c, w in scored if n == best]
    return [c for c, _ in hits], (hits[0][1] if len(hits) == 1 else [])


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", default=SRC)
    ap.add_argument("--check", action="store_true")
    args = ap.parse_args()

    if not os.path.exists(args.src):
        sys.exit(f"Missing {args.src}")
    ws = openpyxl.load_workbook(args.src, data_only=True).worksheets[0]

    fdr, relabelled, problems = {}, [], []
    for label, games in read_rows(ws):
        hits, wrong = identify(games)
        if len(hits) != 1:
            problems.append(f"row {label!r}: matched {len(hits)} clubs ({hits or 'none'})")
            continue
        code = hits[0]
        for md, sheet_opp, real_opp in wrong:
            problems.append(f"{code} MD{md}: workbook lists {sheet_opp}, the fixture is "
                            f"{real_opp} — difficulty colour kept, opponent text ignored")
        if CELL_ABBR[label] != code:
            relabelled.append(f"{label} -> {code} ({ts.TEAM_NAMES[code]}); the label says "
                              f"{CELL_ABBR[label]} ({ts.TEAM_NAMES[CELL_ABBR[label]]})")
        for md, (_, _, v) in enumerate(games, start=1):
            fdr.setdefault(str(md), {})[code] = v

    clubs = {c for m in fdr.values() for c in m}
    print(f"parsed {len(clubs)} clubs x {len(fdr)} matchdays")
    if relabelled:
        print("\nrow labels corrected from the fixtures (source workbook has these wrong):")
        for r in relabelled:
            print(f"  {r}")
    if problems:
        print("\nsource-data problems (handled, not fatal):")
        for p in problems:
            print(f"  {p}")
    missing = sorted(set(ts.TEAM_NAMES) - clubs)
    if missing:
        print(f"\nno FDR for {len(missing)} club(s): {missing}")

    # Agreement with the model's OWN FDR. Read the untouched model baseline from
    # the projections file rather than team_stats — a previous import may already
    # have merged workbook values there, which would compare the data to itself.
    baseline = {}
    if os.path.exists(PROJ):
        baseline = (json.load(open(PROJ)) or {}).get("fdr_model", {})
    same = adj = tot = 0
    for md_s, vals in fdr.items():
        for code, v in vals.items():
            mine = baseline.get(md_s, {}).get(code)
            if mine is None:
                continue
            tot += 1
            same += (mine == v)
            adj += abs(mine - v) <= 1
    if tot:
        print(f"\nagreement with the Elo-derived baseline: exact {same/tot:.0%}, "
              f"within one band {adj/tot:.0%}  ({tot} club-matchdays)")
    else:
        print("\nno model baseline to compare against — rerun build_md_projections.py")

    if args.check:
        return

    json.dump({"fdr": fdr, "source": os.path.basename(args.src),
               "colour_legend": COLOUR_TO_FDR}, open(OUT, "w"), indent=2)
    print(f"\nSaved → {OUT}")

    # Fold into the per-matchday projections the app reads.
    if os.path.exists(PROJ):
        blob = json.load(open(PROJ))
        for md_s, vals in fdr.items():
            blob.setdefault("fdr", {}).setdefault(md_s, {}).update(vals)
            blob.setdefault("fdr_source", {})[md_s] = "workbook"
        json.dump(blob, open(PROJ, "w"), indent=2)
        print(f"Merged into → {PROJ}")


if __name__ == "__main__":
    main()
